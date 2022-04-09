from openpyxl import load_workbook


class getDataFromExcel:


    def getdataexcel(self):

        # to start the excel or load the excel
        book = load_workbook("C:\\Users\\Lokesh\\Desktop\\Python\\pythonProject1\\excelDemo.xlsx")
        # to get the active sheet
        sheet = book.active
        # to get the access to the sheet and print first row and column value
        cell = sheet.cell(row=1, column=1)
        # to print the value
        print(cell.value)
        Dict={}
        for row in range(1, sheet.max_row + 1):
            if sheet.cell(row=row, column=1).value == "Testcase2":
                for col in range(1, sheet.max_column + 1):
                    Dict[sheet.cell(row=1, column=col).value] = sheet.cell(row=row, column=col).value
        print(Dict)