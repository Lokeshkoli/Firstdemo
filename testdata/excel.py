import openpyxl
from openpyxl import load_workbook
#to start the excel or load the excel
book = load_workbook("C:\\Users\\Lokesh\\Desktop\\Python\\pythonProject1\\excelDemo.xlsx")
#to get the active sheet
sheet= book.active
#to get the access to the sheet and print first row and column value
cell= sheet.cell(row=1,column=1)
#to print the value
print(cell.value)

#to write the value
sheet.cell(row=2, column=2).value="LokeshKoli"
print(sheet.cell(row=2, column=2).value)

#to get total row count
print(sheet.max_row)

#to get total coluum count
print(sheet.max_column)

#to print all the first row value
for row in range(1,sheet.max_row+1):
    print(sheet.cell(row=row,column=1).value)

# to print all the excel value
print("All the value")
for row in range(1,sheet.max_row+1):
    for col in range(1,sheet.max_column+1):

        print(sheet.cell(row=row, column=col).value)

print("second row value")
# to get value for specifc row
for row in range(1, sheet.max_row+1):
    if sheet.cell(row=row, column=1).value=="Testcase2":
        for col in range(2, sheet.max_column+1):
            print(sheet.cell(row=row, column=col).value)

print("pass in dictonary")
Dict={}

for row in range(1, sheet.max_row+1):
    if sheet.cell(row=row, column=1).value == "Testcase2":
        for col in range(1, sheet.max_column+1):
            Dict[sheet.cell(row=1, column=col).value] = sheet.cell(row=row, column=col).value
print(Dict)

